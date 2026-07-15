"""
Advanced Excel parser that handles multiple sheets, data types, and large files.
Uses Polars for high-performance reading.
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Any

import polars as pl
from fastapi import UploadFile

from app.core.constants import MAX_ROWS_PER_SHEET

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Parses Excel workbooks into Polars DataFrames.
    Supports .xlsx, .xls, .xlsm, and .csv files.
    """

    @staticmethod
    async def parse(file: UploadFile) -> dict[str, pl.DataFrame]:
        """
        Parse an uploaded file into a dict of sheet_name -> DataFrame.
        For CSV files, a single "Sheet1" entry is returned.
        """
        file_bytes = await file.read()
        file_ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename else "xlsx"

        if file_ext == "csv":
            return await ExcelParser._parse_csv(file_bytes)
        return await ExcelParser._parse_excel(file_bytes)

    @staticmethod
    async def _parse_excel(file_bytes: bytes) -> dict[str, pl.DataFrame]:
        """Parse Excel binary into sheet dict."""
        try:
            workbook = pl.read_excel(
                BytesIO(file_bytes),
                sheet_id=0,  # 0 = all sheets
            )
        except Exception as e:
            logger.warning("Polars read_excel failed, trying openpyxl fallback: %s", e)
            return await ExcelParser._fallback_parse(file_bytes)

        sheets: dict[str, pl.DataFrame] = {}

        if isinstance(workbook, dict):
            for sheet_name, df in workbook.items():
                sheets[sheet_name] = ExcelParser._clean_dataframe(df, sheet_name)
        else:
            sheets["Sheet1"] = ExcelParser._clean_dataframe(workbook, "Sheet1")

        return sheets

    @staticmethod
    async def _parse_csv(file_bytes: bytes) -> dict[str, pl.DataFrame]:
        """Parse CSV into a single sheet."""
        try:
            df = pl.read_csv(BytesIO(file_bytes), infer_schema_length=10000)
        except Exception as e:
            logger.warning("Polars CSV parse failed, trying pandas fallback: %s", e)
            import pandas as pd
            pdf = pd.read_csv(BytesIO(file_bytes))
            df = pl.from_pandas(pdf)
        return {"Sheet1": ExcelParser._clean_dataframe(df, "Sheet1")}

    @staticmethod
    async def _fallback_parse(file_bytes: bytes) -> dict[str, pl.DataFrame]:
        """Fallback using openpyxl when Polars can't read the file."""
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        sheets: dict[str, pl.DataFrame] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = []
            for row in rows[1:]:
                data_rows.append([v if v is not None else None for v in row])
            df = pl.DataFrame(data_rows, schema=headers, orient="row")
            sheets[sheet_name] = ExcelParser._clean_dataframe(df, sheet_name)
        wb.close()
        return sheets

    @staticmethod
    def _clean_dataframe(df: pl.DataFrame, sheet_name: str) -> pl.DataFrame:
        """Clean and normalize a DataFrame."""
        if df.height == 0:
            return df

        # Truncate if too large
        if df.height > MAX_ROWS_PER_SHEET:
            logger.warning(
                "Sheet '%s' has %d rows, truncating to %d",
                sheet_name, df.height, MAX_ROWS_PER_SHEET,
            )
            df = df.head(MAX_ROWS_PER_SHEET)

        # Drop fully empty columns
        df = df.drop([c for c in df.columns if df[c].null_count() == df.height])

        # Rename unnamed columns
        new_cols = []
        for i, c in enumerate(df.columns):
            if c.startswith("Unnamed") or c.strip() == "":
                new_cols.append(f"column_{i}")
            else:
                new_cols.append(c)
        if new_cols != df.columns:
            df.columns = new_cols

        return df